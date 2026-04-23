import openpyxl
from django.db import IntegrityError
from .models import Category, Product
from pytils.translit import slugify

def import_from_excel(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    categories_created = 0
    products_created = 0
    errors = []

    if 'category' in wb.sheetnames:
        ws = wb['category']
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            try:
                name = str(row[0]).strip()
                description = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                is_active = True

                cat, created = Category.objects.get_or_create(
                    name=name,
                    defaults={
                        'description': description,
                        'is_active': is_active,
                        'slug': slugify(name)
                    }
                )
                if created:
                    categories_created += 1
            except IntegrityError as e:
                errors.append(f'Строка {row_idx} (category): {e}')

    if 'products' in wb.sheetnames:
        ws = wb['products']
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue
            try:
                name = str(row[0]).strip()
                category_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                description = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                price = float(row[3]) if len(row) > 3 and row[3] else 0
                discount = int(row[4]) if len(row) > 4 and row[4] else 0
                is_active = True
                try:
                    category = Category.objects.get(name=category_name)
                except Category.DoesNotExist:
                    errors.append(f'Строка {row_idx} (products): категория {category_name} не найдена')
                    continue
                product, created = Product.objects.get_or_create(
                    name=name,
                    defaults={
                        'category': category,
                        'description': description,
                        'price': price,
                        'discount_percent': discount,
                        'is_active': is_active,
                        'slug': slugify(name)
                    }
                )
                if created:
                    products_created += 1
                else:
                    product.price = price
                    product.discount_percent = discount
                    product.save()
            except (ValueError, IntegrityError) as e:
                errors.append(f'Строка {row_idx} (products): {e}')

    msg = f'Импорт завершен. Категорий создано: {categories_created}, Товаров создано: {products_created}'
    if errors:
        msg += f'   Ошибки: {'; '.join(errors)}'
    return msg